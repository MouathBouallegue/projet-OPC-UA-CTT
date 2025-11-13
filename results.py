import xml.etree.ElementTree as ET
import openpyxl
import re
from openpyxl.chart import PieChart, Reference
from openpyxl.styles import PatternFill
# Load and parse the XML file
tree = ET.parse('Serv-Config.results.xml')  # Replace with your XML file path
root = tree.getroot()

# List of test steps to include
include_test_steps = [ 
"Address Space Model",
"Address Space Base",
"Address Space from xml",
"Address Space Base",
"Address Space Complex DataType",
"Address Space Events",
"Address Space Method",
"Address Space ObjectType with BaseType",
"Address Space ReferenceType",
"Address Space VariableType",
"Address Space View",
"BaseDataType with BaseType Method",
"Aggregate - AnnotatedValue", "Aggregate - Average", "Aggregate - Count", "Aggregate - Data", "Aggregate - DataError", "Aggregate - DataInvalid", "Aggregate - Duration", "Aggregate - DurationError", "Aggregate - DurationInvalid", "Aggregate - End", "Aggregate - Error", "Aggregate - Event", "Aggregate - Annotation", "Aggregate - AnnotationError", "Aggregate - AnnotationInvalid", "Aggregate - Hertzian", "Aggregate - HertzianTime", "Aggregate - HertzianTimeError", "Aggregate - HertzianTimeInvalid", "Aggregate - Minimum", "Aggregate - Minimum2", "Aggregate - MinimumActualTime", "Aggregate - MinimumActualTimeError", "Aggregate - MinimumActualTimeInvalid", "Aggregate - MinimumTime", "Aggregate - MinimumTimeError", "Aggregate - MinimumTimeInvalid", "Aggregate - NumberOfChanges", "Aggregate - PercentGood", "Aggregate - PercentBad", "Aggregate - Range", "Aggregate - RangeError", "Aggregate - RangeInvalid", "Aggregate - StandardDeviation", "Aggregate - StandardDeviationSample", "Aggregate - Start", "Aggregate - StatusCode", "Aggregate - TimeAverage", "Aggregate - TimeAverageError", "Aggregate - TimeAverageInvalid", "Aggregate - Total", "Aggregate - TotalError", "Aggregate - TotalInvalid", "Aggregate - Variance", "Aggregate - VarianceSample", "Aggregate - WorstQuality", "Aggregate - WorstQuality2","Event C Acknowledge", "Event C Alarm", "Event C Base", "Event C Condition", "Event C ConditionBranch", "Event C Confirm", "Event C Discrete", "Event C Enable", "Event C Event", "Event C ExclusiveDeviation", "Event C ExclusiveLevel", "Event C ExclusiveLimit", "Event C ExclusiveRateOfChange", "Event C NonExclusiveDeviation", "Event C NonExclusiveLevel", "Event C NonExclusiveLimit", "Event C NonExclusiveRateOfChange", "Event C OffNormal", "Event C Refresh", "Event C RefreshEnd", "Event C RefreshStart", "Event C Suppress", "Event C SystemOffNormal","Attribute - Alternate AccessLevel", "Attribute - Read", "Attribute - Read Complex", "Attribute - Write Simple", "Attribute - Write Value", "Attribute - Write DataValue & Timestamps", "Attribute - Write Mask","Base Info Care System", "Base Info Connection System", "Base Info Engineering Unit", "Base Info EventNotifierType", "Base Info GlobalReferenceMethod", "Base Info ModelChange", "Base Info OptionSet", "Base Info RecommendedRendering", "Base Info ProgressEventType", "Base Info PropertyChange", "Base Info SemanticDescription", "Base Info ServerCapabilities", "Base Info State System", "Base Info State System - substates existing system", "Base Info TopologySystem", "Base Info UnitOfMeasure","Data Access - ArrayItem", "Data Access - AxisInformation", "Data Access - DataItem", "Data Access - MultiDimensionalArrayItem", "Data Access - NetworkItem", "Data Access - Range", "Data Access - SemanticChange", "Data Access - ToolItem","Discovery Accept Registration", "Discovery Accept Registration Security", "Discovery Configuration", "Discovery FindServers Filter", "Discovery FindServers Self", "Discovery GetEndpoints", "Discovery Register", "Discovery Register2","Historical Access Data Value Node Read AtTime", "Historical Access Data Value Read", "Historical Access Insert Data", "Historical Access Modified Values", "Historical Access Raw Modify", "Historical Access Server Capabilities", "Historical Access Update Value","Monitored Items - CreateMonitoredItems", "Monitored Items - DeleteMonitoredItems", "Monitored Items - ModifyMonitoredItems", "Monitored Items - SetMonitoringMode", "Monitored Items - SetTriggering", "Monitored Items - CreateSubscription", "Monitored Items - DeleteSubscriptions", "Monitored Items - ModifySubscription", "Monitored Items - Publish", "Monitored Items - ReprioritizeSubscription", "Monitored Items - SetPublishingMode", "Monitored Items - TransferSubscriptions", "Monitored Items - Read", "Monitored Items - Sampling", "Monitored Items - ValueChange","Node Management Add Node", "Node Management Add Ref", "Node Management Browse", "Node Management BrowseNext", "Node Management Delete Node", "Node Management Delete Ref","ByObjectType - AggregateFunction", "ByObjectType - Annotation", "ByObjectType - AuditConditionEventType", "ByObjectType - AuditEventType", "ByObjectType - AuditHistoryEvent", "ByObjectType - AuditUpdateMethodEventType", "ByObjectType - AuditWriteUpdateEventType", "ByObjectType - BaseDataVariableType", "ByObjectType - BaseEventType", "ByObjectType - BaseObjectType", "ByObjectType - DiscreteItemType", "ByObjectType - PropertyType", "ByObjectType - ReferenceType", "ByObjectType - Structure", "UAFI Asset Component Analog Base", "UAFI Asset Component Discrete Base", "UAFI Asset Component Discrete Static", "UAFI Asset Component Dynamics Base", "UAFI Asset Component Dynamics Static", "UAFI Asset Component Event Base", "UAFI Asset Component Event Static", "UAFI Asset Component Function Base", "UAFI Asset Component Function Static", "UAFI Asset Component Information Base", "UAFI Asset Component Information Static", "UAFI Asset Component Parameter Base", "UAFI Asset Component Parameter Static", "UAFI Asset Server Base", "UAFI Automation Component Communication Network", "UAFI Automation Component Control Network", "UAFI Automation Component Execution Unit", "UAFI Automation Component Field Device", "UAFI Automation Component Function Block", "UAFI Automation Component Process Unit", "UAFI Configuration Support History", "UAFI Configuration Support", "UAFI Connector Analog Input", "UAFI Connector Analog Output", "UAFI Connector Base", "UAFI Connector Discrete Input", "UAFI Connector Discrete Output", "UAFI Connector Information", "UAFI Connector Parameter", "UAFI Connector Signal", "UAFI Control Group Base", "UAFI Control Group Methods", "UAFI Functional Entity Alarm", "UAFI Functional Entity Base", "UAFI Functional Entity Controller", "UAFI Functional Entity Conveyor", "UAFI Functional Entity Drive", "UAFI Functional Entity FlowController", "UAFI Functional Entity Furnace", "UAFI Functional Entity HandlingDevice", "UAFI Functional Entity HeatExchanger", "UAFI Functional Entity Machine", "UAFI Functional Entity MaterialStorage", "UAFI Functional Entity MixingDevice", "UAFI Functional Entity PackagingDevice", "UAFI Functional Entity Parameter", "UAFI Functional Entity Pipe", "UAFI Functional Entity Plant", "UAFI Functional Entity PowerSupply", "UAFI Functional Entity ProcessChamber", "UAFI Functional Entity ProcessModule", "UAFI Functional Entity Pump", "UAFI Functional Entity Robot", "UAFI Functional Entity Sensor", "UAFI Functional Entity Tank", "UAFI Functional Entity Valve", "UAFI Functional Entity VisionSystem", "UAFI Functional Entity Workstation","Protocol Configuration", "Protocol Soap Binary", "Protocol Soap Binary WsSecurity", "Protocol Soap Xml", "Protocol Soap Xml WsSecurity","Security Role Based Application", "Security Administration", "Security Administration - XML Schema", "Security Basic 256 Sha256", "Security Basic 256", "Security Basic 128Rsa15", "Security Basic 256Sha256", "Security Certificate Administration", "Security Certificate Revocation", "Security Default Application Instance", "Security Encryption Required", "Security - No Application Authentication", "Security None", "Security  Role Based Authorization", "Security Row Level Authorization", "Security Signing Required", "Security Time Based - Configuration", "Security Time Based - No Revoke Support", "Security Time Based - Symmetric", "Security User Authentication", "Security User Role Based Permissions", "Security User Token Based Windows", "Security UserName Password", "Security User X509","PubSub - UA Binary Encoding", "PubSub - UA Binary Unordered", "PubSub - UA Json Fixed DataSet Settings","Redundancy Server", "Redundancy Server Transparent","Session Activate", "Session Close", "Session Create", "Session Delete", "Session Service Diagnostics", "Session Service SetMonitoringMode", "Session Service TransferSubscriptions", "Session Monitored Items Create", "Session Monitored Items Delete", "Session Monitored Items Modify", "Session Monitored Items SetMonitoringMode", "Session Monitored Items SetTriggering","Session Base", "Session Cancel", "Session Change User", "Session General Service Behaviour", "Session Minimum 1", "Session Minimum 2 Parallel", "Session Minimum 10 Parallel", "Session Minimum 50 Parallel", "Session Minimum 500 Parallel","Subscription Basic", "Subscription Durable", "Subscription Minimum 1", "Subscription Minimum 02", "Subscription Minimum 05", "Subscription Minimum 10", "Subscription Publish Discard Policy", "Subscription Publish Min 02", "Subscription Publish Min 05", "Subscription Publish Min 10", "Subscription Transfer","View Basic", "View Minimum Continuation Point 01", "View Minimum Continuation Point 05", "View Minimum Continuation Point 10", "View Minimum Continuation Point 50", "View RegisterNodes", "View TranslateBrowsePath","UserDefinedCU","UserDefinedCG","Base Eventing","Monitor Events","Address Space Model", "Aggregates", "Alarms and Conditions", "Attribute Services", "Auditing", "Base Information", "Data Access", "Discovery Services", "History", "Method Services", "Monitored Item Services", "Node Management Services", "PubSub General", "Protocol and Encoding", "Redundancy", "Security", "Session Services", "Subscription Services", "View Services",
"Security None",
"Security User Name Password",
"Security Administration - XML Schema",
"Security User X509",
"Security Certificate Validation",
"Security Certificate Administration",
"Security None CreateSession ActivateSession",
"Security None CreateSession ActivateSession 1.01",
"Security Basic 128Rsa15",
"Security Basic 256Sha256",
"Security Aes128-Sha256-RsaOaep",
"Security Aes256-Sha256-RsaPss",
"Security Administration",
"Security - No Application Authentication",
"Security Basic 256",
"Security User Anonymous",
"Security User IssuedToken Kerberos",
"Security User IssuedToken Kerberos Windows",
"Security Encryption Required",
"Security Signing Required",
"Security Time Synch - Configuration",
"Security Time Synch - NTP/OS Based Support",
"Security Time Synch - UA Based Support",
"Security Default ApplicationInstanceCertificate",
"Security No Application Authentication",
"001.js","002.js","003.js","004.js","005.js","006.js","007.js","008.js","009.js","0010.js","0011.js","0012.js","0013.js","0014.js","0015.js","0016.js","0017.js","0018.js","0019.js","0020.js","0021.js","0022.js","001-01.js", "001-02.js", "001-03.js", "001-04.js", "002-01.js", "002-02.js", "002-03.js", "002-04.js", "003-01.js", "003-02.js", "003-03.js", "003-04.js", "004-01.js", "004-02.js", "004-03.js", "004-04.js", "005-01.js", "005-02.js", "005-03.js", "Err-004.js", "Err-005.js", "Err-006.js", "Err-007.js", "Err-008.js", "Err-009.js", "Err-010.js", "Err-011.js", "Err-012.js", "Err-013.js", "Err-014.js", "Err-015.js", "Err-016.js", "Err-017.js", "Err-018.js", "Err-019.js", "Err-020.js", "Err-021.js", "Err-022.js", "Err-023.js", "Err-024.js", "Err-025.js","test.js","Err-026.js", "Err-027.js", "Err-028.js", "Err-029.js", "Err-030.js", "Err-031.js", "Err-032.js", "Err-033.js", "Err-034.js", "Err-035.js", "Err-036.js", "Err-037.js", "Err-038.js", "Err-039.js", "Err-040.js", "Err-041.js", "Err-042.js", "Err-043.js", "Err-044.js", "Err-045.js", "Err-046.js", "Err-047.js", "Err-048.js", "Err-049.js", "Err-050.js", "Err-051.js", "Err-052.js", "Err-053.js", "Err-054.js", "Err-055.js", "Err-056.js", "Err-057.js", "Err-058.js", "Err-059.js", "Err-060.js",

"000.js", "001.js", "002.js", "003.js", "004.js", "005.js", "006.js", "007.js", "008.js", "009.js", "010.js", "011.js", "012.js", "013.js", "014.js", "015.js", "016.js", "017.js", "018.js", "019.js", "020.js", "021.js", "022.js", "023.js", "024.js", "025.js", "026.js", "027.js", "028.js", "029.js", "030.js", "031.js", "032.js", "033.js", "034.js", "035.js", "036.js", "037.js", "038.js", "039.js", "040.js", "041.js", "042.js", "043.js", "044.js", "045.js", "046.js", "047.js", "048.js", "049.js", "050.js", "051.js", "052.js", "053.js", "054.js", "055.js", "056.js", "057.js", "058.js", "059.js", "060.js", "061.js", "062.js", "063.js", "064.js", "065.js", "066.js", "067.js", "068.js", "069.js", "070.js", "071.js", "072.js", "073.js", "074.js", "075.js", "076.js", "077.js", "078.js", "079.js", "080.js", "081.js", "082.js", "083.js", "084.js", "085.js", "086.js", "087.js", "088.js", "089.js", "090.js", "091.js", "092.js", "093.js", "094.js", "095.js", "096.js", "097.js", "098.js", "099.js", "100.js",

"AASET-001.js", "AASET-002.js", "AASET-004.js", "AASET-005.js", "ACSET-001.js", "ACSET-002.js", "ACSET-004.js", "AOSCET-001.js", "AOSCET-002.js", "AOSCET-003.js", "AOSCET-004.js", "AOSCET-005.js", "AOSCET-006.js", "ASET-003.js",


"Test_001.js", "Test_002.js", "Test_003.js", "Test_004.js", "Test_005.js", "Test_006.js", "Test_007.js", "Test_008.js", "Test_009.js", "Test_010.js", "Test_011.js", "Test_012.js", "Test_013.js", "Test_014.js", "Test_015.js", "Test_016.js", "Test_017.js", "Test_018.js", "Test_019.js", "Test_020.js", "Test_021.js", "Test_022.js", "Test_023.js", "Test_024.js", "Test_025.js", "Test_026.js", "Test_027.js", "Test_028.js", "Test_029.js", "Test_030.js", "Test_031.js", "Test_032.js", "Test_033.js", "Test_034.js", "Test_035.js", "Test_036.js", "Test_037.js", "Test_038.js", "Test_039.js", "Test_040.js", "Test_041.js", "Test_042.js", "Test_043.js", "Test_044.js", "Test_045.js", "Test_046.js", "Test_047.js", "Test_048.js", "Test_049.js", "Test_050.js", "Test_051.js", "Test_052.js", "Test_053.js", "Test_054.js", "Test_055.js", "Test_056.js", "Test_057.js", "Test_058.js", "Test_059.js", "Test_060.js", "Test_061.js", "Test_062.js", "Test_063.js", "Test_064.js", "Test_065.js", "Test_066.js", "Test_067.js", "Test_068.js", "Test_069.js", "Test_070.js", "Test_071.js", "Test_072.js", "Test_073.js", "Test_074.js", "Test_075.js", "Test_076.js", "Test_077.js", "Test_078.js", "Test_079.js", "Test_080.js", "Test_081.js", "Test_082.js", "Test_083.js", "Test_084.js", "Test_085.js", "Test_086.js", "Test_087.js", "Test_088.js", "Test_089.js", "Test_090.js", "Test_091.js", "Test_092.js", "Test_093.js", "Test_094.js", "Test_095.js", "Test_096.js", "Test_097.js", "Test_098.js", "Test_099.js", "Test_100.js",

"Address Space Atomicity",
"Address Space Base",
"Address Space Events",
"Address Space Complex Datatypes",
"Address Space Method",
"Address Space UserWriteMask Multilevel",
"Address Space User Level Security Base",
"Address Space UserWriteAccess",
"Data Access",
"Data Access AnalogItemType",
"Data Access DataItems",
"Data Access MultiState",
"Data Access PercentDeadBand",
"Data Access Semantic Changes",
"Data Access TwoState",
"Data Access ArrayItemType",
"Data Access ComplexNumber",
"Data Access DoubleComplex Number"
"Address Space WriteMask",
"insecureChannelWithCertificatesAndNonce.js",
"insecureChannelWithNonceNoCertificates.js",
"insecureChannelWithNonceOnly.js",
"insecureChannelNoCertOrNonceExpectSecure.js",
"provideInvalidCertificateExpectChannelOpen.js",
"doAttemptConsumeChannels.js",
"dosAttackConsumeSecureChannelsCreateSessions.js",

"insecureChannelAuthSuccess.js",
"secureChannelAuthSuccess.js",
"insecureChannelEmptyUserPassExpectTokenInvalid.js",
"insecureChannelNonExistentUserExpectTokenRejected.js",
"insecureChannelValidUserEmptyPassNonceExpectGoodOrRejected.js",
"insecureChannelValidUserEmptyPassNoNonceExpectGoodOrRejected.js",
"insecureChannelWrongPasswordExpectTokenRejected.js",
"insecureChannelUsernameDoesNotExistExpectTokenRejected.js",
"insecureChannelPolicyIdMismatchExpectTokenInvalid.js",
"insecureChannelEncryptionAlgorithmMismatchExpectTokenInvalid.js",
"secureChannelEmptyEncryptionAlgorithmExpectTokenInvalid.js",
"insecureChannelValidUserNoAccessExpectAccessDenied.js",
"getEndpointsVerifyUniqueUserIdentificationPolicyId.js",
"validCertForValidConnection.js",
"validCertForNotTrustedUser.js",
"notYetValidCertTrusted.js",
"expiredX509Certificate.js",
"revokedUserX509Cert.js",
"applicationInstanceCertificateInstead.js",
"specifyDifferentPolicyId.js",
"invalidUserCertificate.js",
"validCertLocalCANotTrusted.js",
"validTrustedCertLocalCATrusted.js",
"validTrustedCertLocalCANotTrustedButKnown.js",
"validNotTrustedCertLocalCANotTrustedButKnown.js",
"notTrustedCertRevokedServerPolicyId.js",
"validTrustedUserCertEmptyUserIdentitySignature.js",
"validTrustedUserCertInvalidUserIdentitySignature.js",
"validTrustedUserCertUserIdentitySignatureInvalidAlgorithm.js",
"SecurityCertificateValidation:ActivateInsecureSessionValidateCertPerSpecification.js",
"SecurityCertificateValidation:ConnectWithClientCertSignedByUntrustedButKnownCAWithoutRevocation.js",
"SecurityCertificateValidation:AttemptSecureSessionWithEmptyClientCertificate.js",
"SecurityCertificateValidation:SecureChannelUntrustedCertificateExpectsBadSecurityChecksFailed.js",
"SecurityCertificateValidation:SecureChannelExpiredCertificateExpectsGoodOrBadCertificateTimeInvalid.js",
"SecurityCertificateValidation:SecureChannelNotYetValidCertificateExpectsGoodOrBadCertificateTimeInvalid.js",
"SecurityCertificateValidation:SecureChannelIssuerUnknownExpectsBadSecurityChecksFailed.js",
"SecurityCertificateValidation:SecureChannelCertificateSignedWithWrongKeyExpectsBadSecurityChecksFailed.js",
"SecurityCertificateValidation:UsingInsecureConnectionSecurityNoneActivateSessionWhileSendingRevokedCertificate.js",
"SecurityCertificateValidation:AttemptSecureChannelAndSendACA.js",
"SecurityCertificateValidation:ConnectWithExpired/UntrustedCertificate.js",
"SecurityCertificateValidation:SecureChannelIssuedCertificateExpectsGood.js",
"SecurityCertificateValidation:EncryptedChannelUsingRevokedCertificate.js",
"SecurityCertificateValidation:ConnectUsingUntrustedIssuedCertOfCAWithNoRevocationList.js",
"SecurityCertificateValidation:ConnectUsingAnUntrustedIssuedCertOfCAWithNoRevocationListAvailable.js",
"SecurityCertificateValidation:ConnectUsingAnUntrustedIssuedCertOfCAThatIsNotTrustedButAvailable.js",
"SecurityCertificateValidation:ConnectUsingAnUntrustedIssuedCertOfCAThatIsNotTrustedButAvailable.js",
"SecurityCertificateValidation:AttemptSecureChannelAndSendNotTrustedCertificateIssuedByUnknownCertificateAuthority.js",
"SecurityCertificateValidation:ConnectUsingARevokedCertificateThatIsNotTrusted.js",
"SecurityCertificateValidation:ConnectUsingATrustedClientCertificate.js",
"SecurityCertificateValidation:ConnectUsingATrustedClientCertificate(sha1-1024).js",
"SecurityCertificateValidation:ConnectUsingATrustedClientCertificate(sha1-2048).js",
"SecurityCertificateValidation:ConnectUsingATrustedClientCertificate(sha2-2048).js",
"SecurityCertificateValidation:ConnectUsingATrustedClientCertificate(sha2-4096).js",
"SecurityCertificateValidation:ConnectUsingATrustedClientCertificate(sha2-2048).js",
"Write_IndexRange_Equals_One_For_Each_Core_DataType.js",
"Write_IndexRange_Equals_OneHalf_For_Each_Core_DataType_As_Array.js",
"Write_IndexRange_Equals_ZeroTwo_To_First_Three_Elements_Of_Array_For_Each_DataType.js",
"Write_IndexRange_Equates_To_Last_Three_Items_Of_An_Array.js",
"Write_Array_Complete_Array_For_Each_DataType.js",
"Write_Element_Outside_Bounds_Of_Array.js",
"Write_IndexRange_One_Index_Of_MultiDimensional_Array_First_Element_From_Second_Dimension.js",
"Write_IndexRange_Range_Of_Indexes_Of_MultiDimensional_Array_First_Element_From_Each_Dimension.js",
"Write_IndexRange_Range_Of_Indexes_Of_MultiDimensional_Array_First_Three_Elements_Of_Each_Dimension.js",
"Write_IndexRange_Range_Of_Indexes_Of_MultiDimensional_Array_Last_Three_Elements_Of_Each_Dimension.js",
"Write_IndexRange_To_Subset_Of_Array_Verify_Timestamps_Updated.js",
"Error_IndexRange_Is_Out_Of_Bounds_Int32_Divide_By_2.js",
"Error_IndexRange_Is_Invalid_7_2_For_Node_DataType.js",
"Error_IndexRange_Syntax_Invalid_Used_For_DataType_Node.js",
"Error_IndexRange_Invalid_Syntax_Range_Not_Actually_Specified_As_Range.js",
"Error_IndexRange_Invalid_Using_Negative_Numbers_For_Each_DataType.js",
"Write_IndexRange_Equals_One_For_Each_Core_DataType.js",
"Write_IndexRange_Equals_OneHalf_For_Each_Core_DataType_As_Array.js",
"Write_IndexRange_Equals_ZeroTwo_To_First_Three_Elements_Of_Array_For_Each_DataType.js",
"Write_IndexRange_Equates_To_Last_Three_Items_Of_An_Array.js",
"Write_Array_Complete_Array_For_Each_DataType.js",
"Write_Element_Outside_Bounds_Of_Array.js",
"Write_IndexRange_One_Index_Of_MultiDimensional_Array_First_Element_From_Second_Dimension.js",
"Write_IndexRange_Range_Of_Indexes_Of_MultiDimensional_Array_First_Element_From_Each_Dimension.js",
"Write_IndexRange_Range_Of_Indexes_Of_MultiDimensional_Array_First_Three_Elements_Of_Each_Dimension.js",
"Write_IndexRange_Range_Of_Indexes_Of_MultiDimensional_Array_Last_Three_Elements_Of_Each_Dimension.js",
"Write_IndexRange_To_Subset_Of_Array_Verify_Timestamps_Updated.js",
"Error_IndexRange_Is_Out_Of_Bounds_Int32_Divide_By_2.js",
"Error_IndexRange_Is_Invalid_7_2_For_Node_DataType.js",
"Error_IndexRange_Syntax_Invalid_Used_For_DataType_Node.js",
"Error_IndexRange_Invalid_Syntax_Range_Not_Actually_Specified_As_Range.js",
"Error_IndexRange_Invalid_Using_Negative_Numbers_For_Each_DataType.js",
"AccessLevelEx_AtomicityFlags_Check.js",
"Default_Atomicity_Behavior_ManualCheck.js",
"Default_Atomicity_Behavior_ManualCheck_2.js",
"BaseNodeClassComplianceCheck.js",
"TypeSystemStructureValidation.js",

"EventNotifierAttributeValidation.js",
"ServerObjectEventReception.js",
"WriteToAccessLevel_CurrentWrite.js",
"InvalidWrite_ReadOnlyAccess.js",
"WriteToWritableAttributes.js",
"WriteDeniedByUserAccessLevel.js",
"UserReadAccess.js",
"AccessLevelRead.js",

"WriteAllWritableAttributesBasedOnWriteMask.js",
"WriteMultipleAttributesWithOneInvalidAttributedId.js",
"WriteValidValuesToWritableAttributesBasedOnWriteMask.js",
"WriteToInvalidNodesMultipleAttributes.js",

"ReadSingleScalarNodeValueWithServerTimestamp.js",
"ReadMultipleAttributesFromValidNode.js",
"readAllAttributesFromVariableNode.js",
"readMultipleAttributesWithMaxAgeZero.js",
"readDataValueWithBothTimestamps.js",
"readDataValueWithSourceTimestampOnly.js",
"readDataValueWithServerTimestampOnly.js",
"readDataValueWithNoTimestamps.js",
"readBrowseNameAttribute.js",
"readBrowseName_MultipleNodes.js",
"ReadAllMandatoryAttributes_MultipleNodeTypes.js",
"ReadSameAttributeMultipleTimes_SingleNode.js",
"ReadEachSupportedScalarDataType.js",
"Read_MaxAge_Exceeds_Int32_With_Both_Timestamps.js",
"Read_SingleNode_SequentialAttributes_NoSourceTimestamp.js",
"Read_ServerState_ExpectRunning.js",
"Read_Array_Attribute_ExpectArraySizeGreaterThanTwo.js",
"Read_MultipleArrayAttributes_ValidateBuiltinTypesAndArrayFlag.js",
"ReadSingleElementFromArrayByIndexRange_ValidateAgainstFullArrayRead.js",
"ReadArraySubrange_IndexRange2to4_CompareWithFullArray.js",
"ReadArray_Last3Elements_IndexRange.js",
"ReadEventNotifierAttribute.js",
"ReadNonValueAttributesWithTimestampCheck.js",
"ReadMultiDimensionalArrayValueAttribute.js",
"ReadMultiDimensionalArrayValues_DifferentDataTypes.js",
"AttributeRead_MultiDimArray_SingleElement.js",
"ReadRange_MultiDimArray_2To4_SecondDimension.js",
"ReadRange_Last3_LastDimension_MultiDimArray.js",
"ReadArray_IndexRange_BoundsCheck_PartialReturn.js",

"Read_InvalidAttribute_Executable_NotSupported_BadAttributeInvalid.js",
"Read_Attributes_MultipleValidOneInvalid_ErrorForInvalid.js",
"Read_SameInvalidAttribute_MultipleRequests_BadAttributeInvalid.js",
"Read_Attribute_FromInvalidNodeId.js",
"Read_BrowseName_EmptyNodeId.js",
"ReadValidAttributeFromNodeIdWithEmptyStringId.js",
"ReadMultipleAttributes_MixedValidUnknownInvalidNodes.js",
"ReadValidAttributes_FromMultipleNonExistentNodes.js",
"ReadValidAttributes_FromNodesWithInvalidNodeIdSyntax.js",
"ReadEmptyNodesArray.js",
"ReadArrayWithInvalidIndexRangeShouldReturnBadIndexRangeNoData.js",
"ReadArrayWithNegativeIndexRangeReturnsBadIndexRangeInvalid.js",
"ReadWithNegativeMaxAgeReturnsBadMaxAgeInvalid.js",
"ReadAttributesWithInvalidIndexRangeShouldReturnBadIndexRangeNoData.js",
"ReadNotReadableNodeShouldReturnBadNotReadableOrBadUserAccessDenied.js",
"ReadValueWithinInvalidDataEncodingShouldReturnBadDataEncodingInvalid.js",
"ReadWithInvalidTimestampsToReturnShouldReturnBadTimestampsToReturnInvalid.js",
"ReadWithinInvalidIndexRangeContainingDashShouldReturnBadIndexRangeInvalid.js",
"ReadWithSingleIndexRangeShouldReturnBadIndexRangeInvalid.js",
"ReadWithBackwardIndexRangeShouldReturnBadIndexRangeInvalid.js",

"Write_IndexRange_Equals_One_For_Each_Core_DataType.js",
"Write_IndexRange_Equals_OneHalf_For_Each_Core_DataType_As_Array.js",
"Write_IndexRange_Equals_ZeroTwo_To_First_Three_Elements_Of_Array_For_Each_DataType.js",
"Write_IndexRange_Equates_To_Last_Three_Items_Of_An_Array.js",
"Write_Array_Complete_Array_For_Each_DataType.js",
"Write_Element_Outside_Bounds_Of_Array.js",
"Write_IndexRange_One_Index_Of_MultiDimensional_Array_First_Element_From_Second_Dimension.js",
"Write_IndexRange_Range_Of_Indexes_Of_MultiDimensional_Array_First_Element_From_Each_Dimension.js",
"Write_IndexRange_Range_Of_Indexes_Of_MultiDimensional_Array_First_Three_Elements_Of_Each_Dimension.js",
"Write_IndexRange_Range_Of_Indexes_Of_MultiDimensional_Array_Last_Three_Elements_Of_Each_Dimension.js",
"Write_IndexRange_To_Subset_Of_Array_Verify_Timestamps_Updated.js",
"Write_IndexRange_Is_Out_Of_Bounds_Int32_Divide_By_2.js",
"Write_IndexRange_Is_Invalid_7_2_For_Node_DataType.js",
"Write_IndexRange_Syntax_Invalid_Used_For_DataType_Node.js",
"Write_IndexRange_Invalid_Syntax_Range_Not_Actually_Specified_As_Range.js",

"Write_IndexRange_Invalid_Using_Negative_Numbers_For_Each_DataType.js",

"Read_Request_Default_Binary_Encoding.js",
"Read_Request_DataEncoding_Of_Xml.js",
"Read_Request_Unknown_Encoding_For_Node.js",
"Read_Request_Encoding_For_Non_Value_Attribute_Of_Variable.js",
"CreateMonitoredItems_Request_Unsupported_DataEncoding_Modbus.js",
"CreateMonitoredItems_DataEncoding_Namespace_Does_Not_Exist.js",
"CreateMonitoredItems_Request_Encoding_For_Non_Value_Attribute.js",

"Write_Valid_Attribute_Of_Node_NoTimestamps.js",
"Write_Value_Attribute_Multiple_Nodes.js",
"Write_to_Single_Node_Multiple_Times_In_Same_Call.js",
"Write_AccessLevel_CurrentWrite_Expects_Good.js",
"Write_Minimum_Value_For_Each_Supported_DataType.js",
"Write_Max_Value_Of_Each_Supported_DataType.js",
"Write_ByteString_To_Byte_Expects_Success.js",
"Write_LocalizedText_All_Some_No_Parameters_Of_Structure_Type.js",
"Write_Values_That_Are_Subtype_Of_UInteger.js",
"Write_Values_That_Are_Subtype_Of_ULongInteger.js",
"Write_NaN_To_All_Floats.js",
"Write_String_In_Extended_Codepage.js",
"Write_Value_Attribute_NoStatus_NoTime_Node_ValueRank_Array_Entire_Array.js",
"Write_Value_Attribute_NoStatus_NoTime_Several_Nodes_ValueRank_Array_Entire_Array.js",
"Write_Value_Attribute_NoStatus_NoTime_Node_ValueRank_MultiDimensional_Array_Entire_MultiDimensional_Array.js",
"Write_Value_Attribute_NoStatus_NoTime_Several_Nodes_ValueRank_MultiDimensional_Array_Entire_MultiDimensional_Array.js",
"Write_Empty_Nodes_To_WriteArray.js",
"Write_To_Invalid_Attribute_Of_Invalid_Node.js",
"Write_Valid_Node_For_Node_Using_Invalid_Syntax.js",
"Write_Value_To_And_DisplayName_Name_Attributes_On_Multiple_Invalid_Nodes.js",

"Write_Invalid_Attribute.js",
"Write_Invalid_Attribute_To_Valid_Node_Multiple_Times_In_Same_Call.js",
"Write_To_Node_Whose_AccessLevel_Is_ReadOnly_Expect_Bad_NotWritable.js",
"Write_Value_Using_Wrong_DataType.js",
"Write_Null_To_Each_DataType.js",
"WriteIntegerType_Receives_Values_Of_Incorrect_SubType.js",
"WriteInteger_Receives_Values_Of_Incorrect_SubType.js",
"WriteInteger_Receives_Values_Of_Incorrect_SubType_AnotherCase.js",
"WriteError_Write_LocalizedText_With_Non_Existent_Locale.js",

"eventFilterWithContradictoryConditionsShouldFailOrReturnNoEvents.js",

"ServerObject.StructureConsistency.NamespaceMapping.js",
"CoreSystemNodes.BrowseName.Read.js",

"Read_ServerDiagnostics_BrowseName.js",
"ServerDiagnostics_Type_DefinedCorrectly.js",
"Check_CurrentSessionCount_And_CumulatedSessionCount.js",
"Check_SamplingIntervalDiagnosticsArray.js",
"Check_SubscriptionDiagnosticsArray.js",
"Check_SubscriptionDiagnosticsArray_In_Real_Time.js",
"Verify_ModifyCount_Subscription_Diagnostic.js",
"Check_RepublishMessageCount_Subscription_Diagnostic.js",
"Check_PublishRequestCount_And_DataChangeNotificationsCount_Subscription_Diagnostics.js",

"Check_DisabledMonitoredItemCount_Diagnostics_Property.js",
"Check_MonitoringQueueOverflowCount_Subscription_Diagnostic.js",
"Check_Current_CumulatedSubscriptionCount.js",
"Repeat_Current_CumulatedSubscriptionCount.js",
"Check_ServerDiagnostics_EnabledFlag.js",
"Verify_EnabledFlag_Toggle_DoesNot_Increase_SessionCount.js",
"Static_Diagnostics_Return_BadNotReadable_When_EnabledFlag_Is_False.js",
"Test_Sessions_Are_Added_To_SessionsDiagnosticsSummary.js",
"Check_Security_RejectedSessionCount.js",
"SessionDiagnostics_RejectedSessionCount.Security.RejectedSessionCount.IncrementVerification.js",
"Call_GetMonitoredItems_With_Two_Subscriptions_Two_Items.js",
"CreateSubscription_With_Zero_Items.js",
"Two_Subscriptions_Both_Checked.js",
"Invalid_SubscriptionId.js",
"Call_GetMonitoredItems_SubscriptionId_Belongs_To_Another_Session.js",

"EventQueueOverflow.SingleEventLostCheck.js",

"Validate_Correctness_Of_ResendData_Method.js",
"Verify_DataChangeNotifications_Sent_AllMonitoredItems_MonitoringModeReporting.js",
"Verify_DataChangeNotifications_Sent_AllMonitoredItems_MonitoringModeReporting_PublishingIntervalNotTriggered.js",
"Verify_DataChangeNotifications_NotSent_AllMonitoredItems_MonitoringModeSampling.js",

"Verify_DataChangeNotifications_NotSent_AllMonitoredItems_MonitoringModeDisabled.js",
"Verify_DataChangeNotifications_Only_Generated_For_MonitoredItems_Of_Subscription_MonitoringModeReporting.js",
"Verify_DataChangeNotifications_Sent_EvenIf_MaxNotificationsPerPublish_Exceeded.js",
"Verify_ResendData_Method_DoesNot_Trigger_Resending_Of_Events.js",
"Verify_Server_Only_Generates_DataChangeNotifications_For_Given_SubscriptionID.js",
"Validate_Server_Keeps_Operating_When_Calling_ResendData_With_Empty_Subscription.js",
"Verify_After_Calling_ResendData_Only_One_Value_Per_MonitoredItem_Reported.js",
"Invoke_ResendData_With_Non_Existent_Subscription.js",
"ResendData_Method_Cannot_Be_Invoked_From_Another_Session.js",
"Call_ResendData_Method_Without_Having_Subscription_In_Server.js",

"Verify_ServerCapabilities_Object.js",
"Check_MinSupportedSampleRate.js",
"Check_MaxBrowseContinuationPoints.js",
"Check_MaxArrayLength.js",
"Check_MaxStringLength.js",
"Check_MaxByteStringLength.js",
"Check_MaxNodesPerRead.js",
"Check_MaxNodesPerWrite.js",
"Check_MaxNodesPerMethodCall.js",
"Check_MaxNodesPerBrowse.js",
"Check_MaxNodesPerRegisterNodes.js",

"CheckTypesFolderStructure.js",
"DataTypesValidation.js",
"EventTypesValidation.js",
"ObjectTypesValidation.js",
"ReferenceTypesValidation.js",

"VariableTypesValidation.js",

"Type_Defined_In_AddressSpace.js",
"DataItems_ShouldOperate_With_TranslateBrowsePathsToNodeIds.js",
"DataItems_ShouldOperate_With_TranslateBrowsePathsToNodeIds_MultipleDataItems.js",
"Read_Value_Attribute_DataItemNodes_EachDataType_Byte_Double_Float_GUID_Int16_Int32_Int64_UInt16_UInt32_UInt64_SByte_String.js",
"Write_Three_Values_EachSupportedDataType_Min_Max_Middle_Of_DataTypeRange.js",
"Write_To_Multiple_DataItemNodes_AllSupportedDataTypes_SingleCall_Max_Value_Of_CorrespondingDataType.js",
"Write_To_Definition_Property_Of_DataItem_AcquireHandle_Using_TranslateBrowsePathsToNodeIds.js",
"Write_To_ValuePrecision_Property_Of_DataItem_AcquireHandle_Using_TranslateBrowsePathsToNodeIds.js",
"Write_To_DataItemNode_ValuePrecision_ReadBack_Values_Same_Or_Else_Differ_Within_Precision.js",
"Write_To_DataItemNode_ValuePrecision_Integer_Value_ReadBack_Values_Same_Or_Else_Differ_Within_Range_Specified_By_Precision.js",
"Write_Value_To_DataItemNode_Float_Double_ThatMatchesIts_ValuePrecision.js",
"Write_Value_To_DataItemNode_Float_Double_ExceedsIts_ValuePrecision.js",
"Write_Current_SystemTime.js_DataItemNode_DataType_DateTime_Supports_ValuePrecision_Variable.js",
"Subscribe_AddMonitoredItems_DataItem_EachSupportedDataType_EnabledSubscription_CallPublish_VerifyValuesMatchExpected.js",
"Subscribe_AddMonitoredItem_EachSupportedDataType_EnabledSubscription_WriteMax_Value_Corresponding_DataType_CallPublish.js",
"Subscribe_AddMonitoredItem_AbsoluteDeadband_Value_10_EachSupportedDataType_EnabledSubscription_CallPublish_Pass_NotPassDeadband.js",
"Browse_AvailableInverseReferences_DataItemNode.js",
"Browse_AvailableReferences_BothDirections_DataItemNode.js",
"TranslateBrowsePathsToNodeIds_GetProperties_For_AllConfiguredItems.js",
"Read_Attributes_Value_DataType_AccessLevel_MinimumSamplingInterval.js",
"Write_Value_Using_WrongDataType_To_DataItem_HasProperty_Definition.js",
"Write_Value_UsingWrongDataType_To_DataItem_HasProperty_ValuePrecision.js",

"MultiStateDiscreteType_StructureValidation.js",
"Read_Node_Of_This_Type_Check_EnumStrings_Variable_Exists.js",
"Read_MultipleNodes_Of_ThisType_InSingleCall.js",
"Write_To_SingleNode_ThisType_EachValue_SupportedByEnumerationLength.js",
"Subscription_SingleNode_ThisType_Loop_Write_CallPublish_CompareReceivedVsWritten",
"Subscription_MultipleNodes_ThisType_WriteValues_CallPublish_CompareReceivedVsWritten",
"Subscription_EnumStrings_Variable",
"Browse_All_Forward_References",
"Browse_Node_Of_TwoStateDiscreteType_Requesting_All_Forward_References",
"Browse_Node_Of_MultiState_Type_Requesting_All_Forward_References",
"TranslateBrowsePathsToNodeIds_Request_EnumValues_And_ValueAsText",
"Write_To_EnumStrings_Property",
"Write_To_ValueAsText_Property",
"Write_Value_Exceeds_Range_Defined_By_EnumStrings_Property",
"Write_Value_As_String_Within_Bounds_Of_EnumStrings_Length",

"TwoStateDiscreteType_Structure",
"WriteMultipleTwoStateDiscreteValues",
"WriteMultipleTwoStateDiscreteValuesFalse",
"WriteMultipleTwoStateDiscreteTypesAlternating",
"subscribeWritePublishCompareLoop",
"subscribeMultipleNodesWritePublishCompareLoop",
"subscribeMultipleNodesWritePublishCompareLoop",
"browseForwardInverseBothReferences",
"translateBrowsePathsToNodeIdsBothProperties",
"WriteToTrueStateAndFalseState",
"WriteTrueBadTypeMismatch",
"WriteTrueUppercaseBadTypeMismatch",
"WriteIncorrectStringDataType",
"WriteIncorrectStringDataType3",
"WriteIncorrectStringDataType2",
"WriteIncorrectStringDataType4",
"WriteIncorrectStringDataType5",
"WriteIncorrectStringDataType6",
"WriteIncorrectStringDataType3",
"WriteIncorrectStringDataType7",
"testObservedTimestampsReturn",
"testStartEndDateVaryingNodesPartialRequest",
"testStartEndDateNoHistoryData",
"testStartEndDateNoHistoryDataBoundaryValues",
"testStartEndDateNoDataBetweenTimeBounds",
"testStartEndDateNoDataBetweenTimestampsVaryingOrderAndBounds",
"testStartEndDateNoDataBetweenButExistsAtEndpoints",
"testContinuationPointsMaintainedUntilEOF",
"testContinuationPointsMaintainedUntilEOF_Loop2",
"testReceiveBoundsWithStartEndDatePartialRequest",
"testIndexRangeParameterAllArrayBasedNodes",
"testIndexRangeParameterExceedsDimensions",
"testTwoDimensionalIndexRangeParameterAllArrayNodes",
"testReceiveBoundsStartEndSingleNodeRequest",
"testReceiveBoundsPartialExistVaryingNodesPartialRequest",
"requestHistoryMax5ValuesExistingDataBoundaryValues",
"requestRangeWithBadQualityData",
"requestRangeWithUncertainQualityData",
"requestHistoryMax5ValuesExistingDataBoundaryValues_2",
"initMultipleHistoryReadRequestsExceedContinuationPoints",
"errSpecifyEmptyInvalidRequestHeader",
"errInvalidTimestampsToReturnRequest",
"errInvalidTimestampsToReturnValue",
"errRequestServerTimestampsToReturnUnsupported",
"errRequestBothTimestamps",
"errDoNotPopulateReadRawModifiedDetails",
"errRequestHistoryInvalidNodeId",
"errRequestHistoryNodeDoesNotExist",
"errRequestIndexRangeNonArrayNode",
"errRequestSyntacticallyIncorrectIndexRangeArrayNode",
"errUnknownEncodingValue",
"errRequestRawValuesNodeNotAccessible",
"errReusePreviouslyUsedContinuationPoint",
"errUseBrowseCPInHistoryCall",
"errReadHistoryReleaseCPThenReuse",
"errRequestUnsupportedTimestampFormat",
"errReadModifiedValuesWithReturnBoundsTrue",
"errReadRawValuesNoRecordsTimespan",
"errReadRawHistoryStartEndDatesNotExist",
"errReadRawHistoryServerNotCollectingData",









"errReadRawDataHistorianOffline",
"errSpecifyIndexRangeNonArrayNode_2",
"errSpecifyRandomContinuationPoint",
"errSpecifyHistoryValidNodeNoHistoryAccess",
"errSpecifyEmptyInvalidRequestHeader_2",
"errDoNotSpecifyTwoOfThreeRequiredParameters",



"basic128Rsa15_SignOnly",
"basic128Rsa15_SignEncrypt",
"basic128Rsa15_DosAttemptConsumeChannels",
"basic128Rsa15_DosAttackConsumeSecureChannelsCreateSessions",


"basic256Sha256_SignOnly",
"basic256Sha256_SignEncrypt",
"basic256Sha256_DosAttemptConsumeChannels",
"basic256Sha256_DosAttackConsumeSecureChannelsCreateSessions",

"aes128Sha256RsaOaep_SignOnly_Success",
"aes128Sha256RsaOaep_SignEncrypt_Success",
"aes128Sha256RsaOaep_DosAttemptConsumeChannels",
"aes128Sha256RsaOaep_DosAttackConsumeSecureChannelsCreateSessions",

"aes256Sha256RsaPss_SignOnly",
"aes256Sha256RsaPss_SignEncrypt",
"aes256Sha256RsaPss_DosAttemptConsumeChannels",
"aes256Sha256RsaPss_DosAttackConsumeSecureChannelsCreateSessions",


"basic256_SignOnly",
"basic256_SignEncrypt",
"basic256_DosAttemptConsumeChannels",
"basic256_DosAttackConsumeSecureChannelsCreateSessions",

"createSubscriptionDefaultParameters",
"createSubscriptionPublishingIntervalOne",
"createSubscriptionPublishingIntervalZero",
"createSubscriptionPublishingIntervalIntMax",
"createSubscriptionLifetimeKeepAliveZero",
"createSubscriptionLifetime3KeepAlive1",
"createSubscriptionLifetimeEqualsKeepAlive",
"createSubscriptionLifetimeLessThanKeepAlive",
"createSubscriptionLifetimeLess Than3xKeepAlive",
"createSubscriptionLifetimeUInt32MaxKeepAliveHalfMax",
"createSubscriptionLifetimeUInt32MaxKeepAliveHalfMax_2",
"createSubscriptionLifetimeKeepAliveUInt32Max",
"createSubscriptionLifetimeKeepAliveBothUInt32Max",
"createSubscriptionPublishingDisabledNoPublish",
"createSubscriptionNoMonitoredItemsExpectKeepAlive",
"createSubscriptionNoEarlyExpiry",
"createSubscriptionZeroMonitoredItemsKeepAlives",
"createSubscriptionNoMonitoredItemsKeepAlives_2",
"createSubscriptionNoMonitoredItemsKeepAliveChecking",
"createSubscriptionPublishingDisabledSequenceNum1",
"createSubscriptionZeroMonitoredItemsVariousWaitKeepAlives",
"createSubscriptionOneMonitoredItemVariousTimingValues",
"modifySubscriptionDefaultParameters",
"modifySubscriptionPublishingIntervalGreater",
"modifySubscriptionPublishingIntervalLess",
"modifySubscriptionPubIntervalInvalidReturnsFalse",
"modifySubscriptionPubIntervalRevisesValue",
"modifySubscriptionPubIntervalZeroRevises",
"modifySubscriptionPubIntervalMaxUInt32",
"modifySubscriptionLifetimeKeepAlive3xConflict",
"modifySubscriptionLifetimeKeepAliveZero",
"modifySubscriptionLifetimeKeepAliveRevisesMin",
"modifySubscriptionLifetimeEqualsKeepAliveRevisesMin",
"modifySubscriptionLifetimeLess ThanKeepAlive",
"modifySubscriptionLifetimeLess Than3xKeepAlive",
"modifySubscriptionLifetimeMaxUInt32KeepAliveThird",
"modifySubscriptionLifetimeHalfUInt32KeepAliveMax",
"modifySubscriptionLifetimeKeepAliveBothMaxUInt32",
"modifySubscriptionMaxNotificationsPerPublish",
"modifySubscriptionMaxNotificationsToTenDiscardOldestTrue",
"republishRequestsOutOfOrder",
"setPublishingModeDisabledEnabledSub",
"setPublishingModeEnabledDisabledSub",
"setPublishingModeEnableAlreadyEnabledSub",
"setPublishingModeDisableAlreadyDisabledSub",
"modifySubscriptionSpecifySameSubFiveTimes",
"modifySubscriptionSpecifySameSubIdMultipleTimes",
"publishDefaultParameters",
"publishAcknowledgeValidSequenceNumber",
"publishAcknowledgeMultipleSeqNumbersSingleSub",
"publishAcknowledgeMultipleSeqNumbersMultipleSubs",
"publishAcknowledgeSequenceNumberZero",
"publishVerifyScanRatesDataChanges",
"publishAcknowledgeValidInvalidSeqNumbers",
"republishDefaultParameters",
"republishRetrievesLastThreeUpdates",
"republishIgnoreInitialDataChangeLongWait",
"republishRequestMissingSeqBetweenAcknowledged",
"deleteSubscriptionsDefaultParameters",
"deleteSubscriptionsNoPreDeleteMonitoredItems",
"republishRequestSequenceNumberFuture",
"createSubscriptionTimeoutExtendsOnServiceCall",
"requestHeaderTimeoutHintSmallerThanKeepAlive",
"createSubscriptionNegativePublishingInterval",
"createSubscriptionMaxSubscriptionsPlusOne",
"createSubscriptionPublishingIntervalNaN",
"modifySubscriptionExpiresAfterLifetime",
"modifySubscriptionInvalidSubscription",
"modifySubscriptionIdZero",
"modifySubscriptionPublishingIntervalNegative",
"modifySubscriptionSpecifyNegativePublishingInterval",
"setPublishingModeDisableInvalidSubscription",
"setPublishingModeEnableInvalidSubscription",
"setPublishingModeEnableNonExistentSubNoSessionSubs",
"publishNoSubscriptionDefined",
"publishExpectBadSubscriptionIdInvalidResult",
"publishAcknowledgeSameSeqNumberTwice",
"publishAcknowledgeMultipleInvalidSequenceNumbers",
"publishAcknowledgeMultipleInvalidSeqNumbersUnknown",
"publishAcknowledgeMultipleInvalidSeqNumbersMultipleValidSubs",
"republishNoSubscriptionCreated",
"republishSubscriptionIdZero",
"republishInvalidSubscriptionId",
"republishRetransmitSequenceNumberZero",
"republishInvalidRetransmitSequenceNumber",
"republishRequestAlreadyAcknowledgedSequenceNumber",
"republishOnDeletedSubscription",
"deleteSubscriptionsInvalidSubscriptionId",
"deleteSubscriptionsAlreadyDeletedSubscription",
"deleteSubscriptionsMultipleAlreadyDeleted",
"deleteSubscriptionsEmptyIdsArray",
"deleteSubscriptionsUntilTooManyOperationsOr1000",
"createMonitoredItemsByteStringArrayIndexOutOfBounds",
"modifySubscriptionWithAnotherSessionSubscriptionId",
"setPublishingModeWithAnotherSessionSubscriptionId",
"SubscriptionDurableMethodExistsAndDefined",
"setSubscriptionDurableDefaultParameters",
"setSubscriptionDurableLifetimeHoursMaxUInt32",
"setSubscriptionDurableLifetimeHoursZero",
"setupDurableSubWithOneMonitoredItem",
"addMonitoredItemsLargeBufferSizeToDurableSub",
"establishDurableSubWithMultipleMonitoredItems",
"createDurableSubAndAddMonitoredItems",
"createDurableSubWithOneOrMoreMonitoredItems",
"createDurableSubWithServerNotifierEventFilter",
"createDurableSubWithZeroMonitoredItems",
"createDurableSubscription",
"createSubShortLifespanParams",
"setSubscriptionDurableNoParameters",
"setSubscriptionDurableSpecifyOneParameter",
"setSubscriptionDurableSpecifyThreeUIntParameters",
"setSubscriptionDurableWrongDataTypesAllInParams",
"setSubscriptionDurableOnExistingSubscription",
"setSubscriptionDurableAnotherSessionId",
"setSubscriptionDurableNonExistentId",
"createSubscriptionAckQueuedSeqNumbersOutOfOrder",
"createSubscriptionCheckInDiagnostics",
"subscriptionPrioritiesEqual",
"oneSessionTwoSubscriptions",
"subscriptionWithHigherPriority",
"subscriptionPriorityWaitForStarve",
"modifyOneSubscriptionPriorityHigher",
"modifyOneSubscriptionPriorityLower",
"modifyFirstSubscriptionPriorityHigherThenLower",
"modifySecondSubscriptionPriorityHigherThenLower",
"disableMultipleSubscriptionsSomeActiveInactive",
"enableMultipleSubscriptionsSomeActiveInactive",
"disableAlreadyDisabledSubscriptions",
"enableAlreadyEnabledSubscriptions",
"enableTwoPreviouslyDisabledSubscriptions",
"disableOneSubscriptionOtherActiveSameSession",
"enableTwoAlreadyEnabledSubscriptionsSameSession",
"deleteMultipleSubscriptions",
"publishMultipleSubscriptionServicedWithoutPrejudice",
"republishTwoSubsAckFirstOnlyRepublishSecond",
"setMonitoringModeDistinguishDifferentSubsWhenDisablingOne",
"twoSubscriptionsOneEventsOneDataChanges",
"disableMixValidInvalidSubscriptionIds",
"enablePublishingInvalidSubscriptionIds",
"enablePublishingValidInvalidSubscriptionIds",
"createSubscriptionFiveSubsVaryingPrioritiesCheckNotifications",
"createSubscriptionFiveSubsPerSessionFiveSessions",
"deleteSubscriptionsMultipleValid",
"setPublishingModeEnableFiveDisabledSubsVerifyPublish",
"setPublishingModeCreateFiveDisableTwoVerifyPublish",
"createSubscriptionThreeSubsVariousPriorities",
"publishMultipleSubscriptionServicedWithoutPrejudice",
"setPublishingModeDisableMultipleInvalidSubs",
"publishAcknowledgeMultipleInvalidSeqNumbersMultipleInvalidSubIds",
"deleteSubscriptionsMultipleInvalid",
"deleteSubscriptionsMultipleSomeValidSomeInvalid",
"deleteSubscriptionsMultipleValidAndAlreadyDeleted",
"createTenSubscriptions",
"deleteSubscriptionsMultipleValid",
"setPublishingModeEnableTenPreviouslyDisabledSubs",
"setPublishingModeTenActiveDisableOdd",
"publishMultipleSubscriptionServicedWithoutPrejudice",
"queueTwoPublishRequestsSingleSession",
"deleteSubscriptionWithTwoOutstandingPublish",
"modifySubscriptionCheckTimingsAfterPubIntervalMod",
"republishMessageRetransmissionQueueSizeExpectedTwo",
"publishTimeoutHintCausesBadTimeout",
"queueMorePublishRequestsThanServerSupport",
"browseReferenceTypeIdNullAllReferences.js",
"browseReferencesForwardDirection.js",
"browseReferencesInverseDirection.js",
"browseReferencesSpecifiedRefTypeAndNOSubtypesMatch.js",
"browseReferencesSpecifiedRefTypeAndSubtypesMatch_02.js",
"browseReferencesMatchingNodeClassMask.js",
"browseMultipleNodesInOneCall.js",
"browseMultipleNodesMixValidInvalidIds.js",
"browseAllReferencesNodeClassMaskZero.js",
"browseReturnOnlyOneRefDescriptionField.js",
"browseReferencesSpecifiedRefTypeNoSubtypesParentMatch.js",
"browseReferencesSpecifiedRefTypeSubtypesParentTypeMatch.js",
"browseReferencesNodeClassView.js",
"browseNonNullRefViewFieldsExceptTypeDefinition.js",
"browseReturnRefDescriptionFieldsByResultMask.js",
"browseRefsRefTypeRecursiveSubtypeGrandparentMatch.js",
"browseReturnDiagnosticInfoByBitMask.js",
"browseReturnDiagnosticsWhenZero.js",
"browseBadNodeInInvalidSyntax.js",
"browseBadNodeIdUnknownNotInAddressSpace.js",
"browseBadBrowseDirectionInvalid.js",
"browseBadViewIdUnknownViewNotExists.js",
"browseBadReferenceTypeIdInvalid.js",
"browseBadNothingToDNodesToBrowseEmpty.js",
"browseBadViewIdUnknownSpecifiedViewNotExists.js",
"browseBadReferenceTypeIdNotRefTypeNode.js",
"browseBadNothingToDNoContinuationPoints.js",
"browseBadSecurityCheckFailedNullAuthToken.js",
"browseBadSecurityCheckFailedAuthTokenNotExists.js",
"browseForMaximumOneReference.js",
"browseNextNextRefReleaseContinuationPoints.js",
"browseNextReturnCPWhenReleaseFalseMoreRefs.js",
"browseNextToLastRefGoodEmptyCP.js",
"browseReturnInverseRefsWhenDirectionInverse.js",
"browseReturnRefsSpecifiedTypeNoSubtypes.js",
"browseReturnRefsSpecifiedTypeWithSubtypes.js",
"browseNextReturnRefsMatchingNodeClassMask.js",
"browseNextReturnResultFieldsByResultMask.js",
"browseNextReturnFewerRefsRequestedOrRemaining.js",
"browseNextReturnViewsWhenNodeClassMaskIncludes.js",
"browseNextReturnDiagnosticsByDiagnosticMask.js",
"browseNextNoDiagnosticInfoWhenMaskZero.js",
"browseNextBadNoContinuationPointsMaxReached.js",
"browseNextBadContinuationPointInvalidPreviousSession.js",
"browseNextBadContinuationPointInvalidAfterRelease.js",
"browseNextBadContinuationPointInvalidAlreadyUsed.js",
"browseNextBadSecurityCheckFailedNullAuthToken.js",
"browseNextBadSecurityCheckFailedAuthTokenNotExists.js",
"browseNextBadInvalidTimestampRequestHeader.js",
"browseReturnFiveContinuationPoints.js",
"browseNextReturnMultipleContinuationPoints.js",
"browseNextCombineRefsFromSeparateBrowseCalls.js",
"NodesRegistration:VerifyReferenceToNodesToRegister[0].js",
"NodesRegistration:Verify5NodeReferences.js",
"NodesRegistration:Verify25NodeReferences.js",
"NodesRegistration:Verify50NodeReferences.js",
"NodesRegistration:Verify100NodeReferences.js",
"NodesRegistration:Verify20DuplicateNodeRegistrations.js",
"NodesRegistration:VerifyMixedExistingandNon-ExistentNodeRegistrations.js",
"NodesRegistration:VerifyDiagnosticInfoBasedonBitmask.js",
"NodesRegistration:VerifyNoDiagnosticInfoWhenBitmaskis0.js",
"NodesUnregistration:VerifySuccessfulSingleNodeUnregistration.js",
"NodesUnregistration:VerifySuccessful5NodeUnregistrations.js",
"NodesUnregistration:VerifySuccessful25NodeUnregistrations.js",
"NodesUnregistration:VerifySuccessful50NodeUnregistrations.js",
"NodesUnregistration:VerifySuccessful100NodeUnregistrations.js",
"NodesUnregistration:VerifySuccessfulMixedRegisteredandUnregisteredNodeUnregistrations.js",
"NodesUnregistration:VerifySuccessfulMixedRegisteredandNon-ExistentNodeUnregistrations.js",
"NodesUnregistration:VerifySuccessfulUnregistrationOfNon-ExistentNode.js",
"NodesUnregistration:VerifySuccessfulUnregistrationOfAlreadyUnregisteredNode.js",
"NodesUnregistration:VerifySuccessfulUnregistrationOfMultipleUnregisteredNodes.js",
"NodesUnregistration:VerifyNoUnregistrationWhenSourceNodeisUnregistered.js",
"NodesUnregistration:VerifyDiagnosticInfoBasedonBitmask.js",
"NodesUnregistration:VerifyNoDiagnosticInfoWhenBitmaskis0.js",
"NodesRegistration:HandleEmptyNodesToRegisterArray(Bad_NothingToDo).js",
"NodesRegistration:HandleExcessiveNodeRequests.js",
"NodesRegistration:HandleNon-ExistentSingleNodeId(Bad_NodeIdUnknown).js",
"NodesRegistration:HandleMultipleNon-ExistentNodeId(Bad_NodeIdUnknown).js",
"NodesRegistration:Error-InvalidSingleNodeId(Bad_NodeIdInvalid).js",
"NodesRegistration:Error-MixedValidandInvalidNodes(Bad_NodeIdInvalid).js",
"NodesRegistration:Error-NullAuthenticationToken(Bad_SecurityChecksFailed).js",
"NodesRegistration:Error-MissingAuthenticationToken(Bad_SecurityChecksFailed).js",
"NodesRegistration:Error-InvalidRequestTimestamp(Bad_InvalidTimestamp).js",
"NodesRegistration:Error-NoNodestoUnregister(Bad_NothingToDo).js",
"NodesUnregistration:Error-NullAuthenticationToken(Bad_SecurityChecksFailed).js",
"NodesUnregistration:Error-MissingAuthenticationToken(Bad_SecurityChecksFailed).js",
"NodesUnregistration:Error-InvalidRequestTimestamp(Bad_InvalidTimestamp).js",



"publishFiveConcurrentRequestsSingleSession.js",
"createHalfSupportedSessions.js",
"republishBasic.js",
"publishTimeoutHintCausesBadTimeout.js",
"Err-001.js",
"10SubscriptionsTenPublishCalls.js",
"setPublishingModeCreateTenDisableFiveVerifyPublish.js",
"publishTenConcurrentRequestsQueued.js",
"publishMoreRequestsThanServerHandle.js",
"closeSessionDeleteSubsFalseVerifyNotDeleted.js",
"closeSessionCreateSubsCheckDeleted.js",
"modifySubscriptionTransferedToAnotherSession.js",
"setPublishingModeModifyTransferredSubPublishing.js",
"setPublishingModeModifyMultipleTransferredSubsPublishing.js",
"publishOnTransferredSubscription.js",
"republishOnTransferredSubscription.js",
"republishCloseExistingSessionWithoutKillingSub.js",
"transferSubscriptionToAnotherSessionPublishBoth.js",
"transferSubscriptionToAnotherSessionPublishBoth_02.js",
"transferSubscriptionToAnotherSessionPublishBoth_03.js",
"transferSubscriptionToAnotherSessionPublishBoth_04.js",
"transferSubscriptionToAnotherSessionPublishBoth_05.js",
"transferSubscriptionToAnotherSessionPublishBoth_06.js",
"transferSubscriptionToAnotherSessionPublishBoth_07.js",
"transferSubscriptionToAnotherSession_Simple.js",
"deleteSubscriptionsTransferredSub.js",
"deleteSubscriptionsTransferredSubsAllDeleted.js",
"setMonitoringModeTransferSubCallFromOriginalSession.js",
"deleteSubscriptionsSomeValidInvalidTransferred.js",
"callTransferSubscriptionNoIds.js",
"callTransferSubscriptionMoreIdsThanServerCanHandle.js",
"callTransferSubscriptionInvalidNonExistentId.js",
"createSubMonitoredItemReplicateClientTransferToNewSession.js",
"createSessionUserCredsTransferSub.js",
"createSessionDifferentUserCredsTransferSub.js",
"createSessionAnonymousCredsTransferSubDifferentSecureChannel.js",

"TranslateBrowsePathToNodes:VerifyValidTargetIdFor1PathElement.js",
"TranslateBrowsePathToNodes:VerifyNodeIdofLastBrowsePathElement.js",
"TranslateBrowsePathToNodes:VerifyNodeIdofLastBrowsePathElementWhen10BrowsePathElements.js",
"TranslateBrowsePathToNodes:VerifyNodeIdofLastBrowsePathElementsWhenInverseIsTrue.js",
"TranslateBrowsePathToNodes:VerifyNodeIdofRelativePathElement.ReferenceTypeIdisParentTypeandIncludeSubtypesTrue.js",
"TranslateBrowsePathToNodes:VerifyNodeIdofMatchingBrowseNameWhenReferenceTypeIdisANullNodeId.js",
"TranslateBrowsePathToNodes:Verify4TargetIds,1forEachof4BrowsePathElements.js",
"TranslateBrowsePathToNodes:Verify4TargetIds,1forEachof10BrowsePathElements.js",
"TranslateBrowsePathToNodes:VerifyTargetNodeIdWhenRelativePathElement.ReferenceTypeIdisTargetandIncludeSubtypesTrue.js",
"TranslateBrowsePathToNodes:VerifyTargetNodeIdWhenRelativePathElement.ReferenceTypeIdisGrandparentTypeandIncludeSubtypesTrue.js",
"TranslateBrowsePathToNodes:VerifyDiagnosticInfoasSpecifiedbyReturnDiagnosticBitMask.js",
"TranslateBrowsePathToNodes:VerifyNoDiagnosticInfoWhenReturnDiagnosticsIs0.js",
"TranslateBrowsePathToNodes:Error-InvalidStartingNodeId(Bad_NodeIdInvalid).js",
"TranslateBrowsePathToNodes:Error-GoodMatchforBrowsePathandBadforNonMatchingBrowsePaths.js",
"TranslateBrowsePathToNodes:Error-EmptyRequest(Bad_NothingToDo).js",
"TranslateBrowsePathToNodes:Error-UnknownStartingNodeId(Bad_NodeIdUnknown).js",
"TranslateBrowsePathToNodes:Error-NothingToDoWhenRelativePathisEmpty.js",
"TranslateBrowsePathToNodes:Error-BrowseNameDoesNotExistMatch.js",
"TranslateBrowsePathToNodes:Error-TargetNameisANullQualifiedName.js",
"TranslateBrowsePathToNodes:Error-BadBrowseNameIdWhenBrowseNameisANullBrowseName.js",
"TranslateBrowsePathToNodes:Error-ReferenceTypeIdDoesNotExistMatch.js",
"TranslateBrowsePathToNodes:Error-ContainsInvalidNodeIdforReferenceTypeId.js",
"TranslateBrowsePathToNodes:Error-Non-ExistentNodeIdWhenReferenceTypeIdisANodeIdThatDoesNotExist.js",
"TranslateBrowsePathToNodes:Error-BrowseNameisTooLongorContainsInvalidCharacters.js",
"TranslateBrowsePathToNodes:Error-NoMatchWhenRelativePathElement.ReferenceTypeIdisParentOfReferenceTypeandIncludeSubtypesFalse.js",
"TranslateBrowsePathToNodes:Error-NoMatchWhenReferenceTypeIdisNodeIdofSomethingOtherThanReferenceType.js",
"TranslateBrowsePathToNodes:Error-NoMatchWhenRelativePathElement.IsInverseIsTrueandBrowseNameisInForwardDirection.js",
"TranslateBrowsePathToNodes:Error-AnyMatchAnyNodeWhenReferenceTypeIdisNullandIncludeSubtypesFalse.js",
"TranslateBrowsePathToNodes:TestServerCapabilitiesMaxOperationsPerMethod.js",
"TranslateBrowsePathToNodes:Error-SecurityChecksFailedWhenAuthenticationTokenIsNull.js",
"TranslateBrowsePathToNodes:Error-SecurityChecksFailedWhenAuthenticationTokenDoesNotExist.js",
"TranslateBrowsePathToNodes:Error-BadInvalidTimestampWhenRequestHeader.TimestampIsTooFarOutOfRange.js"














]

# Get the first sentence, and add second/third if needed
def get_first_sentence(text):
    sentences = re.findall(r'[^.?!]+[.?!]', text.strip())
    if not sentences:
        return text.strip()

    first = sentences[0].strip()
    if len(first) >= 25:
        return first

    # Include the second sentence if available
    if len(sentences) > 1:
        second = sentences[1].strip()
        combined = f"{first} {second}"

        # Include the third if second is too short
        if len(second) < 15 and len(sentences) > 2:
            third = sentences[2].strip()
            combined += f" {third}"
        return combined

    return first

# Determine reason from Skipped, Error or Not Supported
def get_reason(node):
    # Check if there is a 'Skipped' child
    skipped = node.find(".//ResultNode[@name='Skipped']")
    if skipped is not None:
        return skipped.get('description', '')

    # Check for 'Error'
    error = node.find(".//ResultNode[@name='Error']")
    if error is not None:
        desc = error.get('description', '')
        if desc.startswith("ActivateSession") or desc.startswith("Write()") or desc.startswith("Read()"):
            return desc
        return get_first_sentence(desc)

    # Check for 'Not Supported'
    not_supported = node.find(".//ResultNode[@name='Not Supported']")
    if not_supported is not None:
        return not_supported.get('description', '')

    return ''

# Process result nodes
def process_results(root):
    results = []

    for result_node in root.findall(".//ResultNode"):
        name = result_node.get('name')
        testresult = result_node.get('testresult')
        description = result_node.get('description', '')

        if name in include_test_steps:
            # Set failure reason
            if testresult == '3':
                reason = "Not Implemented"
            else:
                reason = get_reason(result_node)

            # Classify the test result
            if testresult in ['0', '3', '4', '5']:
                status = 'FAILED'
            elif testresult == '1':
                status = 'WARNING'
            elif testresult == '6':
                status = 'PASSED'
            else:
                continue

            results.append({
                'test_step': name,
                'description': description,
                'result': status,
                'failure_reason': reason if status in ['FAILED', 'WARNING'] else ''
            })

    return results


# Générer les statistiques globales
def generate_summary_stats(results):
    summary = {'PASSED': 0, 'FAILED': 0, 'WARNING': 0}
    for r in results:
        if r['result'] in summary:
            summary[r['result']] += 1
    total = sum(summary.values())
    stats = []
    for key in ['PASSED', 'FAILED', 'WARNING']:
        count = summary[key]
        percent = round((count / total * 100), 2) if total > 0 else 0
        stats.append([key, count, f"{percent}%"])
    return stats

# Exporter vers Excel avec mise en forme et camembert
def export_to_xlsx(results, filename='filtered_test_results.xlsx'):
    wb = openpyxl.Workbook()

    # Feuille 1 : Résultats détaillés
    ws1 = wb.active
    ws1.title = "Test Results"
    ws1.append(['Test Step', 'Description', 'Result', 'Failure Reason'])

    for r in results:
        ws1.append([r['test_step'], r['description'], r['result'], r['failure_reason']])

    # Mise en couleur conditionnelle
    red_fill = PatternFill(start_color="FFC7CE", fill_type="solid")     # FAILED
    yellow_fill = PatternFill(start_color="FFEB9C", fill_type="solid")  # WARNING
    green_fill = PatternFill(start_color="C6EFCE", fill_type="solid")   # PASSED

    for row in ws1.iter_rows(min_row=2, max_col=4):
        result_cell = row[2]
        if result_cell.value == 'FAILED':
            for cell in row:
                cell.fill = red_fill
        elif result_cell.value == 'WARNING':
            for cell in row:
                cell.fill = yellow_fill
        elif result_cell.value == 'PASSED':
            for cell in row:
                cell.fill = green_fill          


      

    # Feuille 2 : Statistiques
    ws2 = wb.create_sheet(title="Statistiques")
    ws2.append(['Statut', 'Nombre', 'Pourcentage'])
    stats = generate_summary_stats(results)
    for row in stats:
        ws2.append(row)

    # === Ajout de la mise en couleur conditionnelle pour Statistiques ===
    green_fill = PatternFill(start_color="C6EFCE", fill_type="solid")    # PASSED
    red_fill = PatternFill(start_color="FFC7CE", fill_type="solid")      # FAILED
    yellow_fill = PatternFill(start_color="FFEB9C", fill_type="solid")   # WARNING

    for row in ws2.iter_rows(min_row=2, max_col=3): # type: ignore
       status_cell = row[0]
       if status_cell.value == 'PASSED':
            for cell in row:
                cell.fill = green_fill
       elif status_cell.value == 'FAILED':
               for cell in row:
                   cell.fill = red_fill
       elif status_cell.value == 'WARNING':
              for cell in row:
                  cell.fill = yellow_fill

            
     

    # Ajouter un graphique camembert
    chart = PieChart()
    chart.title = "Répartition des résultats"
    data = Reference(ws2, min_col=2, max_col=2, min_row=2, max_row=4)
    labels = Reference(ws2, min_col=1, min_row=2, max_row=4)
    chart.add_data(data, titles_from_data=False)
    chart.set_categories(labels)
    ws2.add_chart(chart, "E2")  # Position du graphique

    # Enregistrer le fichier
    wb.save(filename)
    print(f"Résultats enregistrés dans {filename}")


    # (reste du script inchangé)

# Générer un rapport HTML simple pour Jenkins
def export_to_html(results, filename='reports/report.html'):
    import os

    os.makedirs('reports', exist_ok=True)

    with open(filename, 'w', encoding='utf-8') as f:
        f.write('<html><head><meta charset="utf-8"><title>Rapport CTT</title></head><body>')
        f.write('<h1>Rapport de test OPC UA CTT</h1>')

        passed = sum(1 for r in results if r['result'] == 'PASSED')
        failed = sum(1 for r in results if r['result'] == 'FAILED')
        warning = sum(1 for r in results if r['result'] == 'WARNING')

        f.write(f'<p><strong>PASSED :</strong> {passed}<br>')
        f.write(f'<strong>FAILED :</strong> {failed}<br>')
        f.write(f'<strong>WARNING :</strong> {warning}</p>')

        f.write('<table border="1" cellpadding="4" cellspacing="0">')
        f.write('<tr><th>Test Step</th><th>Description</th><th>Result</th><th>Failure Reason</th></tr>')

        for r in results:
            color = {
                'PASSED': '#d4edda',  
                'FAILED': '#f8d7da',
                'WARNING': '#fff3cd'            
            }.get(r['result'], '#ffffff')

            f.write(f'<tr style="background-color:{color}">')
            f.write(f"<td>{r['test_step']}</td>")
            f.write(f"<td>{r['description']}</td>")
            f.write(f"<td>{r['result']}</td>")
            f.write(f"<td>{r['failure_reason']}</td>")
            f.write('</tr>')

        f.write('</table>')
        f.write('</body></html>')

    print(f" Rapport HTML généré : {filename}")


# --- Exécution principale ---
results = process_results(root)
export_to_xlsx(results)
export_to_html(results)


